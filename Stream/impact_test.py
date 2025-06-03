# impact_test.py
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
import io
import os

EXCEL_FILE = "impact_test_data.xlsx"

def save_data_to_excel(data: dict, img_file):
    df = pd.DataFrame([data])
    if not os.path.exists(EXCEL_FILE):
        df.to_excel(EXCEL_FILE, sheet_name="Impact Data", index=False)
    else:
        with pd.ExcelWriter(EXCEL_FILE, mode="a", if_sheet_exists="overlay", engine="openpyxl") as writer:
            reader = pd.read_excel(EXCEL_FILE, sheet_name="Impact Data")
            df = pd.concat([reader, df], ignore_index=True)
            df.to_excel(writer, sheet_name="Impact Data", index=False)

    if img_file:
        wb = load_workbook(EXCEL_FILE)
        sheet_name = f"Image_{data['Test Request No']}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        image = Image.open(img_file)
        buf = io.BytesIO()
        image.save(buf, format="PNG")
        img_bytes = buf.getvalue()
        img = XLImage(io.BytesIO(img_bytes))
        img.width = 500
        img.height = 300
        ws.add_image(img, "A1")
        wb.save(EXCEL_FILE)

def impact_test_page():
    st.header("Impact Test Form")

    with st.form("impact_form"):
        col1, col2 = st.columns(2)
        with col1:
            part_no = st.text_input("Part No.")
            part_name = st.text_input("Part Name")
            model = st.text_input("Model")
            requested_by = st.text_input("Requested By")
        with col2:
            test_request_no = st.text_input("Test Request No")
            test_standard = st.text_input("Test Standard")
            test_start_date = st.date_input("Test Start Date")
            test_name = st.text_input("Test Name")

        st.subheader("Rim Details")
        rim_make = st.text_input("Rim Make")
        rim_size = st.text_input("Rim Size")
        rim_type = st.text_input("Rim Type")
        rim_radius = st.text_input("Static Wheel Radius (mm)")

        st.subheader("Tyre Details")
        tyre_make = st.text_input("Tyre Make")
        tyre_size = st.text_input("Tyre Size")
        max_load = st.text_input("Max Load Capacity of Tyre (Kg)")
        tyre_pressure = st.text_input("Tyre Air Pressure (KPa)")

        st.subheader("Impact Conditions")
        front = st.checkbox("Front")
        rear = st.checkbox("Rear")
        impact_load = st.text_input("Impact Load (Kg)")
        impact_height = st.text_input("Impact Height (mm)")
        striker_type = st.selectbox("Impact Test Striker", ["Standard", "Step"])
        location = st.multiselect("Impact Locations", ["On Spoke", "Between Spoke", "On Valve Hole"])

        image = st.file_uploader("Upload Impact Test Image", type=["png", "jpg", "jpeg"])
        submit = st.form_submit_button("Submit")

        if submit:
            data = {
                "Part No.": part_no,
                "Part Name": part_name,
                "Model": model,
                "Requested By": requested_by,
                "Test Request No": test_request_no,
                "Test Standard": test_standard,
                "Test Start Date": test_start_date,
                "Test Name": test_name,
                "Rim Make": rim_make,
                "Rim Size": rim_size,
                "Rim Type": rim_type,
                "Static Wheel Radius": rim_radius,
                "Tyre Make": tyre_make,
                "Tyre Size": tyre_size,
                "Tyre Max Load": max_load,
                "Tyre Pressure": tyre_pressure,
                "Front": front,
                "Rear": rear,
                "Impact Load": impact_load,
                "Impact Height": impact_height,
                "Striker Type": striker_type,
                "Locations": ", ".join(location)
            }

            save_data_to_excel(data, image)
            st.success("Data and image saved successfully!")
